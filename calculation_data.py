from openpyxl.styles import Alignment, Font, Color
from openpyxl.styles.borders import Border, Side
from span import Span
from support import Support

# стили выравнивания и шрифт ячеек
alignment = Alignment(horizontal="center", vertical="center")
font_big = Font(name='Microsoft Sans Serif', size=8)
font_small = Font(name='Microsoft Sans Serif', size=7)

left_border = Border(left=Side(style='thin', color=Color("969696")))
right_border = Border(right=Side(style='thin', color=Color("969696")))
top_border = Border(top=Side(style='thin', color=Color("969696")))
bottom_border = Border(bottom=Side(style='thin', color=Color("969696")))

col_names = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q',
             'R', 'S', 'T', 'U',
             'V', 'W', 'X', 'Y', 'Z', 'AA']
col_width = [4.43, 6.86, 7.71, 5.57, 6.86, 7.14, 8.43, 7.14, 7, 7.57, 8.43, 7, 8.71, 8.43, 8.43,
             8.43, 8.43, 8.43,
             8.43, 8.14, 8.43, 8.43, 11.29, 9.43, 8.43, 6.57, 9.43]


def end_flag(support, arr):
    if support in arr:
        return True
    else:
        return False


class CalculationData:
    """Подготовленные данные, которые могут провести расчёт"""

    def __init__(self, data):
        self.spans = []
        self.get_spans(data)
        self.supports = []
        self.get_supports(data)
        self.wb = data.workbook
        self.wsdata = data.worksheet
        self.ws = data.workbook.create_sheet("расчёт", 0)
        self.title = 'ТП-' + str(data.ST_number) + ' Л' + str(data.feeder_number)
        self.new_filename = '{} - расчёт.xlsx'.format(self.title)
        self.set_col_width()
        self.write_title()
        self.write_table()
        self.save_file()

    def get_spans(self, data):
        # создаем список с объектами пролетов
        for mark, span, length, lamp in zip(data.list_of_marks, data.list_of_span_support_numbers,
                                            data.list_of_length, data.list_of_lamps):
            flag = False
            if lamp:
                flag = True
            span = Span(mark, span, length, flag)
            self.spans.append(span)

    def get_supports(self, data):
        # создаём список с объектами опор
        temp_span = None
        for cur_support in data.supports_for_calc:
            for cur_span in self.spans:
                if end_flag(cur_support, data.end_supports_list):
                    if int(cur_span.span_support1_number) == cur_support or int(
                            cur_span.span_support2_number) == cur_support:
                        support = Support(cur_support, cur_span)
                        self.supports.append(support)
                        break
                else:
                    if int(cur_span.span_support2_number) == cur_support:
                        temp_span = cur_span
                        continue
                    elif int(cur_span.span_support1_number) == cur_support:
                        support = Support(cur_support, temp_span, cur_span)
                        self.supports.append(support)
                        break

    # функции форматирования ячеек
    def cell_border(self, row, col, side):
        self.ws.cell(row=row, column=col).border = self.ws.cell(row=row, column=col).border + side

    def set_cell_height(self, row, height):
        rd = self.ws.row_dimensions[row]  # get dimension for row 3
        rd.height = height

    def set_cell_width(self, col, width):
        rd = self.ws.column_dimensions[col]  # get dimension for row 3
        rd.width = width

    def merge_cells(self, row, col, nRow):
        self.ws.merge_cells(start_row=row, start_column=col,
                            end_row=row + nRow, end_column=col)

    def merge_cells_of_support(self, first_row, support):
        offset1 = 0
        offset2 = 0
        if support.span1.lamp is True:
            offset1 = 1
        nRow = 1 + offset1
        if support.span2:
            if support.span2.lamp is True:
                offset2 = 1
            nRow = nRow + 2 + offset2

        self.merge_cells(first_row, 1, nRow)
        self.merge_cells(first_row, 8, nRow)
        self.merge_cells(first_row, 9, nRow)
        self.merge_cells(first_row, 14, nRow)
        self.merge_cells(first_row, 15, nRow)
        self.merge_cells(first_row, 19, nRow)
        self.merge_cells(first_row, 20, nRow)
        self.merge_cells(first_row, 25, nRow)
        self.merge_cells(first_row, 26, nRow)
        self.merge_cells(first_row, 27, nRow)

        for row in range(nRow + 1):
            self.cell_border(first_row + row, 1, left_border)
            self.cell_border(first_row + row, 7, right_border)
            self.cell_border(first_row + row, 9, right_border)
            self.cell_border(first_row + row, 13, right_border)
            self.cell_border(first_row + row, 15, right_border)
            self.cell_border(first_row + row, 18, right_border)
            self.cell_border(first_row + row, 20, right_border)
            self.cell_border(first_row + row, 25, right_border)
            self.cell_border(first_row + row, 26, right_border)
            self.cell_border(first_row + row, 27, right_border)
        for col in range(27):
            self.cell_border(first_row, col + 1, top_border)
            self.cell_border(first_row + nRow, col + 1, bottom_border)

    def merge_cells_of_span(self, first_row, span):
        if span:
            nRow = len(span.linelevels) - 1

        self.merge_cells(first_row, 2, nRow)
        self.merge_cells(first_row, 21, nRow)
        self.merge_cells(first_row, 22, nRow)
        self.merge_cells(first_row, 23, nRow)
        self.merge_cells(first_row, 24, nRow)

        for row in range(nRow + 1):
            self.cell_border(first_row + row, 2, right_border)
            self.cell_border(first_row + row, 21, right_border)
            self.cell_border(first_row + row, 22, right_border)
            self.cell_border(first_row + row, 23, right_border)
            self.cell_border(first_row + row, 24, right_border)

    def formatCell(self, r, c, v):
        self.ws.cell(row=r, column=c, value=v).alignment = alignment
        self.ws.cell(row=r, column=c, height=9).font = font_big

    # функции вписывания данных в лист Excel
    def write_line_levels(self, row, line_level):
        self.set_cell_height(row, 9)
        self.ws.cell(row=row, column=3, value=line_level.wire_mark).alignment = alignment
        self.ws.cell(row=row, column=3).font = font_small
        self.ws.cell(row=row, column=4, value=line_level.number_of_wires).alignment = alignment
        self.ws.cell(row=row, column=4).font = font_small
        self.ws.cell(row=row, column=5, value=line_level.diameter).alignment = alignment
        self.ws.cell(row=row, column=5).font = font_small
        self.ws.cell(row=row, column=6, value=line_level.weight).alignment = alignment
        self.ws.cell(row=row, column=6).font = font_small
        self.ws.cell(row=row, column=7, value=line_level.level_height).alignment = alignment
        self.ws.cell(row=row, column=7).font = font_small

    def write_spans(self, row, span):
        self.ws.cell(row=row, column=2, value=span.length).alignment = alignment
        self.ws.cell(row=row, column=2).font = font_big

        def fill_cells(linelevel):
            self.ws.cell(row=row + linelevel, column=10,
                         value=span.Gc[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=10).font = font_small
            self.ws.cell(row=row + linelevel, column=11,
                         value=span.Pm[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=11).font = font_small
            self.ws.cell(row=row + linelevel, column=12,
                         value=span.Gmp[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=12).font = font_small
            self.ws.cell(row=row + linelevel, column=13,
                         value=span.Qm[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=13).font = font_small
            self.ws.cell(row=row + linelevel, column=16,
                         value=span.M12[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=16).font = font_small
            self.ws.cell(row=row + linelevel, column=17,
                         value=span.M13[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=17).font = font_small
            self.ws.cell(row=row + linelevel, column=18,
                         value=span.M134[linelevel]).alignment = alignment
            self.ws.cell(row=row + linelevel, column=18).font = font_small

        for linelevel in range(len(span.linelevels)):
            fill_cells(linelevel)

        number_of_linelevels = len(span.linelevels)
        y = number_of_linelevels - 1
        while y > 0:
            self.write_line_levels(row, span.linelevels[y])
            y -= 1
            row += 1
        self.write_line_levels(row, span.linelevels[0])
        return row

    def write_supports(self, row, support):
        self.ws.cell(row=row, column=1, value=support.number).alignment = alignment
        self.ws.cell(row=row, column=1).font = font_big
        self.ws.cell(row=row, column=8, value=support.number_of_lamps).alignment = alignment
        self.ws.cell(row=row, column=8).font = font_big
        self.ws.cell(row=row, column=9, value=support.lamp_height).alignment = alignment
        self.ws.cell(row=row, column=9).font = font_big
        self.ws.cell(row=row, column=14, value=support.Wm).alignment = alignment
        self.ws.cell(row=row, column=14).font = font_big
        self.ws.cell(row=row, column=15, value=support.Wg).alignment = alignment
        self.ws.cell(row=row, column=15).font = font_big
        self.ws.cell(row=row, column=19, value=support.MWm).alignment = alignment
        self.ws.cell(row=row, column=19).font = font_big
        self.ws.cell(row=row, column=20, value=support.MWm).alignment = alignment
        self.ws.cell(row=row, column=20).font = font_big
        self.ws.cell(row=row, column=21, value=support.m1[0]).alignment = alignment
        self.ws.cell(row=row, column=21).font = font_big
        self.ws.cell(row=row, column=22, value=support.m2[0]).alignment = alignment
        self.ws.cell(row=row, column=22).font = font_big
        self.ws.cell(row=row, column=23, value=support.m3[0]).alignment = alignment
        self.ws.cell(row=row, column=23).font = font_big
        self.ws.cell(row=row, column=24, value=support.maxMoments[0]).alignment = alignment
        self.ws.cell(row=row, column=24).font = font_big
        if support.span2 is not None:
            dRow = len(support.span1.linelevels)
            self.ws.cell(row=row + dRow, column=21, value=support.m1[1]).alignment = alignment
            self.ws.cell(row=row + dRow, column=21).font = font_big
            self.ws.cell(row=row + dRow, column=22, value=support.m2[1]).alignment = alignment
            self.ws.cell(row=row + dRow, column=22).font = font_big
            self.ws.cell(row=row + dRow, column=23, value=support.m3[1]).alignment = alignment
            self.ws.cell(row=row + dRow, column=23).font = font_big
            self.ws.cell(row=row + dRow, column=24,
                         value=support.maxMoments[1]).alignment = alignment
            self.ws.cell(row=row + dRow, column=24).font = font_big

        self.ws.cell(row=row, column=25, value=support.maxMoment).alignment = alignment
        self.ws.cell(row=row, column=25).font = font_big
        self.ws.cell(row=row, column=26, value=support.supportType).alignment = alignment
        self.ws.cell(row=row, column=26).font = font_big
        self.ws.cell(row=row, column=27, value=support.loadCapacity).alignment = alignment
        self.ws.cell(row=row, column=27).font = font_big

    def set_col_width(self):
        # формируем ширину столбцов
        table_width = dict(zip(col_names, col_width))
        for key, value in table_width.items():
            self.set_cell_width(key, value + 0.72)

    def write_title(self):
        # вписываем в ячейки Excel
        self.ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=27)
        self.ws['A1'] = self.title
        self.ws['A1'].font = Font(name='Microsoft Sans Serif', size=11)

    def write_table(self):
        # делаем цикл по опорам
        # заполняем таблицу с этого номера
        cur_row = 2

        for support in self.supports:
            # делаем расчеты
            for linelevel in support.span1.linelevels:
                support.span1.calculate_parameters(linelevel)
            if support.span2:
                for linelevel in support.span2.linelevels:
                    support.span1.calculate_parameters(linelevel)
            support.calculate_parameters()

            # вписываем
            self.merge_cells_of_support(cur_row, support)
            self.write_supports(cur_row, support)
            self.merge_cells_of_span(cur_row, support.span1)
            cur_row = self.write_spans(cur_row, support.span1)
            if support.span2:
                self.merge_cells_of_span(cur_row + 1, support.span2)
                cur_row = self.write_spans(cur_row + 1, support.span2)
            cur_row += 1

        for col in range(27):
            self.cell_border(cur_row, col + 1, top_border)

    def save_file(self):
        self.wb.remove(self.wsdata)
        self.wb.save(self.new_filename)
