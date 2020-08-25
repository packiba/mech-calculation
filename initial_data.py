import re

from openpyxl import load_workbook


class InitialData:
    """Исходные данные для расчётов. Класс создаётся на основе excel файла."""

    def __init__(self, filename):
        self.filename = filename
        self.workbook = None
        self.worksheet = None
        self.ST_number = None
        self.feeder_number = None
        self.supports_for_calc = []
        self.supports_with_lamp = []
        self.total_number_of_spans = None
        self.length = 0.0
        self.read_data_from_file()
        self.initial_data_table = self.clean_unnecessary_data(self.make_table())
        self.list_of_span_support_numbers = []
        self.list_of_supports = []
        self.end_supports_list = []
        self.list_of_marks = []
        self.list_of_length = []
        self.list_of_lamps = []
        self.fill_lists()

    @staticmethod
    def supports_list(span):  # span = '33 - 45' => ['33', '45']
        return re.findall(r'\d+', span)

    @staticmethod
    def make_list(spans_list):
        temp = 0
        end_index_list = []
        for sup1_number, sup2_number in spans_list:
            if temp != 0 and temp != sup1_number:
                end_index_list.append(int(temp))
            if temp != sup1_number:
                end_index_list.append(int(sup1_number))
            temp = sup2_number
        end_index_list.append(int(temp))
        return (end_index_list)

    def cellValue(self, target_row, target_col):
        return self.worksheet.cell(row=target_row, column=target_col).value

    def is_span_included_in_list(self, span, support_list):  # span = '33 - 45'
        span_support_numbers = self.supports_list(span)  # ['33', '45']
        if len(span_support_numbers) == 0:
            return False
        if len(span_support_numbers) == 1:
            number_of_sup = int(span_support_numbers[0])
            if number_of_sup in support_list:
                return True
        if len(span_support_numbers) == 2:
            number_of_sup1 = int(span_support_numbers[0])
            number_of_sup2 = int(span_support_numbers[1])
            if number_of_sup1 in support_list and number_of_sup2 in support_list:
                return True
        return False

    def read_data_from_file(self):
        # открываем файл и берем текущий лист
        full_filename = self.filename + '.xlsx'
        self.workbook = load_workbook(full_filename)
        self.worksheet = self.workbook.active

        # номер КТП (ST - substation transformer)
        self.ST_number = int(re.findall(r'\d+', self.worksheet['A2'].value)[0])
        # номер линии
        self.feeder_number = int(re.findall(r'\d+', self.worksheet['B2'].value)[0])

        col_C = self.worksheet['C']  # опоры пролётов '1 - 2'
        col_H = self.worksheet['H']  # опоры для расчёта (на которых подвес ОКСН)
        col_I = self.worksheet['I']  # опоры, на которых фонарка

        self.supports_for_calc = list(
            item.value for item in col_H if item.value)  # очищаем от пустых ячеек
        self.supports_with_lamp = list(
            item.value for item in col_I if item.value)  # очищаем от пустых ячеек
        self.total_number_of_spans = len(col_C)

    def make_table(self):
        table = []
        for line_number in range(1, self.total_number_of_spans):
            new_row = []
            for col_number in range(3, 6):
                new_row.append(self.cellValue(line_number + 1, col_number))
            span = new_row[0]
            if self.is_span_included_in_list(span, self.supports_with_lamp):
                new_row.append(span)
            else:
                new_row.append('')
            table.append(new_row)
        return table

    def clean_unnecessary_data(self, table):
        new_table = []
        for line in table:
            span = line[0]
            if self.is_span_included_in_list(span, self.supports_for_calc):
                new_table.append(line)
                self.length += float(line[2])
                print(line)
        return new_table

    def fill_lists(self):
        # список пролётов
        for item in self.initial_data_table:
            self.list_of_span_support_numbers.append(self.supports_list(item[0]))

        self.end_supports_list = self.make_list(self.list_of_span_support_numbers)

        for item in self.initial_data_table:
            self.list_of_marks.append(item[1])
            self.list_of_length.append(round(item[2]))
            self.list_of_lamps.append(item[3])
